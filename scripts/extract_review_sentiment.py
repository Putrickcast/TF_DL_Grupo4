"""Build the frontend review-sentiment artifact from the NLP workbook.

The Angular app should consume compact JSON rather than parse Excel in the
browser. This script reads the sentiment workbook produced by the notebook and
writes public/data/review-sentiment.json for the LLM/NLP card and late fusion.
"""

from __future__ import annotations

import json
import math
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT_XLSX = ROOT / "docs" / "llm-rag" / "Cuadro_Control_Analisis_Sentimiento_G4_FINAL.xlsx"
OUTPUT_JSON = ROOT / "public" / "data" / "review-sentiment.json"


def pct(counter: Counter[str], key: str, total: int) -> float:
    return round((counter.get(key, 0) / total) if total else 0, 4)


def build_review_sentiment() -> dict[str, Any]:
    wb = load_workbook(INPUT_XLSX, data_only=True, read_only=True)
    comments = wb["Comentarios_Analizados"]
    headers = [cell.value for cell in next(comments.iter_rows(min_row=1, max_row=1))]
    idx = {header: index for index, header in enumerate(headers)}

    by_id: defaultdict[str, dict[str, Any]] = defaultdict(
        lambda: {"sentiment": Counter(), "emotion": Counter(), "scores": [], "reviews": 0}
    )

    for row in comments.iter_rows(min_row=2, values_only=True):
        listing_id = row[idx["ID Airbnb"]]
        if not listing_id:
            continue
        listing_id = str(listing_id).strip()
        entry = by_id[listing_id]
        entry["reviews"] += 1
        entry["sentiment"][row[idx["sentimiento_3"]] or "Sin dato"] += 1
        entry["emotion"][row[idx["emocion"]] or "Sin dato"] += 1
        score = row[idx["score_sentimiento"]]
        if isinstance(score, (int, float)):
            entry["scores"].append(float(score))

    absa: defaultdict[str, defaultdict[str, dict[str, Any]]] = defaultdict(
        lambda: defaultdict(lambda: {"sentiment": Counter(), "mentions": 0})
    )
    if "Detalle_ABSA" in wb.sheetnames:
        detail = wb["Detalle_ABSA"]
        detail_headers = [cell.value for cell in next(detail.iter_rows(min_row=1, max_row=1))]
        detail_idx = {header: index for index, header in enumerate(detail_headers)}
        for row in detail.iter_rows(min_row=2, values_only=True):
            listing_id = row[detail_idx["ID Airbnb"]]
            aspect = row[detail_idx["Aspecto"]]
            sentiment = row[detail_idx["Sentimiento"]]
            if not listing_id or not aspect or not sentiment:
                continue
            item = absa[str(listing_id).strip()][str(aspect).strip()]
            item["mentions"] += 1
            item["sentiment"][str(sentiment).strip()] += 1

    listings: dict[str, Any] = {}
    for listing_id, entry in sorted(by_id.items()):
        total = entry["reviews"] or 1
        positive_pct = pct(entry["sentiment"], "Positivo", total)
        neutral_pct = pct(entry["sentiment"], "Neutral", total)
        negative_pct = pct(entry["sentiment"], "Negativo", total)
        score = round((positive_pct + 0.5 * neutral_pct) * 100, 1)
        confidence = round(min(100, math.log10(total + 1) / math.log10(60) * 100), 1)
        average_raw_score = (
            round(sum(entry["scores"]) / len(entry["scores"]), 2) if entry["scores"] else None
        )

        top_emotion, top_emotion_count = ("Sin dato", 0)
        if entry["emotion"]:
            top_emotion, top_emotion_count = entry["emotion"].most_common(1)[0]
        emotion_total = sum(entry["emotion"].values()) or 1
        emotions = [
            {"emotion": emotion, "pct": round(count / emotion_total, 4), "count": count}
            for emotion, count in entry["emotion"].most_common()
        ]

        aspects = []
        for aspect, item in absa.get(listing_id, {}).items():
            mentions = item["mentions"] or 1
            aspects.append(
                {
                    "aspect": aspect,
                    "positivePct": pct(item["sentiment"], "Positivo", mentions),
                    "neutralPct": pct(item["sentiment"], "Neutral", mentions),
                    "negativePct": pct(item["sentiment"], "Negativo", mentions),
                    "mentions": item["mentions"],
                }
            )
        aspects.sort(key=lambda item: (item["mentions"], item["positivePct"]), reverse=True)

        listings[listing_id] = {
            "positivePct": positive_pct,
            "neutralPct": neutral_pct,
            "negativePct": negative_pct,
            "reviewCount": entry["reviews"],
            "score": score,
            "confidence": confidence,
            "averageRawScore": average_raw_score,
            "topEmotion": top_emotion,
            "topEmotionPct": round(top_emotion_count / emotion_total, 4),
            "emotions": emotions,
            "aspects": aspects,
        }

    return {
        "meta": {
            "generatedFrom": "docs/llm-rag/Cuadro_Control_Analisis_Sentimiento_G4_FINAL.xlsx",
            "sourceSheets": ["Comentarios_Analizados", "Detalle_ABSA", "Hoja1"],
            "scoreField": "score",
            "scoreMeaning": "Score textual 0-100 para fusion tardia: % positivo + 0.5 * % neutral.",
            "confidenceMeaning": "Confianza 0-100 basada en cobertura de reviews por alojamiento.",
            "listingCount": len(listings),
            "aspectRows": sum(len(item["aspects"]) for item in listings.values()),
        },
        "listings": listings,
    }


def main() -> None:
    payload = build_review_sentiment()
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_JSON}")
    print(f"{payload['meta']['listingCount']} listings, {payload['meta']['aspectRows']} ABSA rows")


if __name__ == "__main__":
    main()
